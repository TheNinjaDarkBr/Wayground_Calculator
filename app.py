import streamlit as st
import pandas as pd
from functools import reduce
import io
from openpyxl.styles import Alignment, PatternFill, Border, Side

st.set_page_config(
    page_title="Meu App",
    layout="wide"  # deixa o app em tela cheia
)

st.title("Carregar Planilha")

# Upload do arquivo
arquivos = st.file_uploader("Selecione os arquivos", type=["xlsx"], accept_multiple_files=True)

dataframes = []

if arquivos:
    for arq in arquivos:
        nome_base = arq.name.split("-")[0]
        df = pd.read_excel(arq, sheet_name='Participant Data')
        df["Name"] = df["First Name"].astype(str) + " " + df["Last Name"].astype(str)
        df = df[['Class Name', "Name", 'Accuracy']]
        df['Accuracy'] = df['Accuracy'].astype(str).str.replace('%','').astype(float)
        df = (
            df.groupby(['Class Name', 'Name'])
                .agg(
                    **{f"Acc-{nome_base}": ('Accuracy', 'max'),
                        f"Tentativa-{nome_base}": ('Accuracy', 'count')}
                )
                .reset_index()
            )
        dataframes.append(df)

    dfFinal = reduce(lambda left, right: pd.merge(left, right, on=["Class Name", "Name"], how="outer"), dataframes)
    dfFinal.fillna(0, inplace=True)

    # Calcular ACC Total (média das médias de acurácia de cada arquivo)
    colunas_acc = [col for col in dfFinal.columns if col.startswith('Acc-')]
    # Calcular a média apenas das colunas de acurácia (ignorando valores NaN)
    dfFinal['ACC Total'] = dfFinal[colunas_acc].mean(axis=1, skipna=True).round(2)
    
    # Input para porcentagem personalizada
    st.write("---")
    st.write("📊 Configurações de Cálculo:")
    porcentagem_input = st.number_input(
        "Digite uma porcentagem para criar uma coluna personalizada (ex: 30 para 30%):",
        min_value=0.0,
        max_value=100.0,
        value=0.0,
        step=0.1,
        help="Se preenchido, será criada uma coluna 'ACC Total por [NÚMERO]%' com o valor do ACC Total multiplicado pela porcentagem"
    )
    
    if porcentagem_input >0:
        # Novo input embaixo do existente
        novo_input = st.number_input(
            "Coloque o valor da media com base na porcentagem selecionada acima (notas acima da media ficarao verdes e abaixo ficarao vermelhas):",
            min_value=0,
            value=0,
            step=1,
            help="Media da porcentagem pra nota acima ou baixa"
        )
    
    # Adicionar coluna personalizada se porcentagem for informada
    if porcentagem_input > 0:
        dfFinal[f'ACC Total por {porcentagem_input}%'] = (dfFinal['ACC Total'] * (porcentagem_input / 100)).round(2)
    
    # Reordenar colunas para colocar ACC Total após Name
    colunas_ordenadas = ['Class Name', 'Name', 'ACC Total']
    if porcentagem_input > 0:
        colunas_ordenadas.append(f'ACC Total por {porcentagem_input}%')
    colunas_restantes = [col for col in dfFinal.columns if col not in colunas_ordenadas]
    dfFinal = dfFinal[colunas_ordenadas + colunas_restantes]

    st.write("📊 Visualização da Planilha Completa:")
    st.dataframe(dfFinal)
    
    # Botão de download com células mescladas
    def to_excel_with_merged_cells(df):
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Criar uma cópia do dataframe para manipulação
            df_export = df.copy()
            
            # Adicionar coluna vazia para mesclar
            df_export.insert(0, 'Class_Name_Merge', '')
            
            # Preencher apenas a primeira linha de cada turma
            current_class = None
            for idx, row in df_export.iterrows():
                if row['Class Name'] != current_class:
                    df_export.at[idx, 'Class_Name_Merge'] = row['Class Name']
                    current_class = row['Class Name']
            
            # Remover a coluna original Class Name
            df_export = df_export.drop('Class Name', axis=1)
            
            # Renomear a coluna mesclada
            df_export = df_export.rename(columns={'Class_Name_Merge': 'Class Name'})
            
            # Escrever no Excel
            df_export.to_excel(writer, index=False, sheet_name='Dados_Consolidados')
            
            # Obter a planilha para mesclar células
            worksheet = writer.sheets['Dados_Consolidados']
            
            # Formatar cabeçalho com cor azul escuro
            for col in range(1, len(df_export.columns) + 1):
                header_cell = worksheet.cell(row=1, column=col)
                header_cell.fill = PatternFill(start_color='405ddb', fill_type='solid')
                header_cell.alignment = Alignment(horizontal='center', vertical='center')
                header_cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
            
            # Lista de cores para as turmas
            cores_turmas = [
                'FFE6E6',  # Rosa claro
                'E6F3FF',  # Azul claro
                'E6FFE6',  # Verde claro
                'FFF0E6',  # Laranja claro
                'F0E6FF',  # Roxo claro
                'FFFFE6',  # Amarelo claro
                'E6FFFF',  # Ciano claro
                'FFE6F0',  # Rosa pálido
                'F0FFE6',  # Verde pálido
                'E6E6FF'   # Azul pálido
            ]
            
            # Criar mapeamento de turmas para cores
            turmas_unicas = df_export['Class Name'].unique()
            turmas_unicas = [t for t in turmas_unicas if t != '']  # Remover strings vazias
            mapeamento_cores = {turma: cores_turmas[i % len(cores_turmas)] for i, turma in enumerate(turmas_unicas)}
            
            # Mesclar células da coluna Class Name
            current_class = None
            start_row = 2  # Começar da linha 2 (após cabeçalho)
            end_row = 2
            
            for idx, row in df_export.iterrows():
                if row['Class Name'] != current_class and row['Class Name'] != '':
                    # Mesclar células da classe anterior se existir
                    if current_class is not None and end_row > start_row:
                        worksheet.merge_cells(f'A{start_row}:A{end_row}')
                        # Centralizar verticalmente e horizontalmente
                        worksheet[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
                        # Aplicar cor da turma
                        if current_class in mapeamento_cores:
                            worksheet[f'A{start_row}'].fill = PatternFill(start_color=mapeamento_cores[current_class], fill_type='solid')
                    
                    # Iniciar nova classe
                    current_class = row['Class Name']
                    start_row = idx + 2  # +2 porque Excel começa em 1 e tem cabeçalho
                    end_row = start_row
                elif row['Class Name'] == current_class:
                    end_row = idx + 2
                else:
                    end_row = idx + 2
            
            # Mesclar última classe
            if current_class is not None and end_row > start_row:
                worksheet.merge_cells(f'A{start_row}:A{end_row}')
                # Centralizar verticalmente e horizontalmente
                worksheet[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
                # Aplicar cor da turma
                if current_class in mapeamento_cores:
                    worksheet[f'A{start_row}'].fill = PatternFill(start_color=mapeamento_cores[current_class], fill_type='solid')
            
            # Definir borda preta
            black_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # Centralizar todas as células de dados e aplicar formatação condicional
            acc_total_col = None
            name_col = None
            acc_personalizada_col = None
            for col_idx, col_name in enumerate(df_export.columns, 1):
                if col_name == 'ACC Total':
                    acc_total_col = col_idx
                elif col_name == 'Name':
                    name_col = col_idx
                elif col_name.startswith('ACC Total por') and col_name.endswith('%'):
                    acc_personalizada_col = col_idx
            
            # Criar mapeamento de linha para turma
            linha_para_turma = {}
            current_class = None
            for idx, row in df_export.iterrows():
                if row['Class Name'] != current_class and row['Class Name'] != '':
                    current_class = row['Class Name']
                linha_para_turma[idx + 2] = current_class  # +2 porque Excel começa em 1 e tem cabeçalho
            
            for row in range(2, len(df_export) + 2):  # Começar da linha 2 (após cabeçalho)
                # Obter a turma da linha atual
                turma_da_linha = linha_para_turma.get(row, '')
                cor_da_turma = mapeamento_cores.get(turma_da_linha, 'FFFFFF')  # Branco como padrão
                
                for col in range(1, len(df_export.columns) + 1):  # Todas as colunas
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Aplicar cor da turma em todas as células da linha (exceto colunas especiais)
                    if col != acc_total_col and col != acc_personalizada_col:
                        cell.fill = PatternFill(start_color=cor_da_turma, fill_type='solid')
                        cell.border = black_border
                    
                    
                    # Aplicar formatação condicional para ACC Total
                    if col == acc_total_col and cell.value is not None:
                        try:
                            acc_value = float(cell.value)
                            if acc_value < 60:
                                cell.fill = PatternFill(start_color='d46161', fill_type='solid')
                                cell.border = black_border
                            if acc_value >= 60:
                                cell.fill = PatternFill(start_color='73c56c', fill_type='solid')
                                cell.border = black_border
                        except (ValueError, TypeError):
                            pass  # Ignorar se não conseguir converter para float
                    
                    # Aplicar formatação condicional para ACC Total por % baseada no novo input
                    if col == acc_personalizada_col and cell.value is not None and novo_input > 0:
                        try:
                            acc_personalizada_value = float(cell.value)
                            if acc_personalizada_value >= novo_input:
                                cell.fill = PatternFill(start_color='73c56c', fill_type='solid')
                                cell.border = black_border
                            else:
                                cell.fill = PatternFill(start_color='d46161', fill_type='solid')
                                cell.border = black_border
                        except (ValueError, TypeError):
                            pass  # Ignorar se não conseguir converter para float
                    

            # Ajustar largura das colunas
            worksheet.column_dimensions['A'].width = 90
            worksheet.column_dimensions['B'].width = 50
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 40
            worksheet.column_dimensions['E'].width = 40 
            worksheet.column_dimensions['F'].width = 40
            worksheet.column_dimensions['G'].width = 40
            worksheet.column_dimensions['H'].width = 40
            worksheet.column_dimensions['I'].width = 40
            worksheet.column_dimensions['J'].width = 40
            worksheet.column_dimensions['K'].width = 40
            worksheet.column_dimensions['L'].width = 40
            worksheet.column_dimensions['M'].width = 40
            worksheet.column_dimensions['N'].width = 40
            worksheet.column_dimensions['O'].width = 40
            worksheet.column_dimensions['P'].width = 40
            worksheet.column_dimensions['Q'].width = 40
            worksheet.column_dimensions['R'].width = 40
            worksheet.column_dimensions['S'].width = 40

        return output.getvalue()
    
    excel_data = to_excel_with_merged_cells(dfFinal)
    st.download_button(
        label="📥 Baixar Planilha Completa (XLSX)",
        data=excel_data,
        file_name="dados_consolidados.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    # Separar por turma
    st.write("---")
    st.write("📚 Visualização por Turma:")
    
    # Obter lista única de turmas
    turmas = dfFinal['Class Name'].unique()
    
    for turma in turmas:
        st.write(f"### Turma: {turma}")
        df_turma = dfFinal[dfFinal['Class Name'] == turma].copy()
        df_turma_display = df_turma.drop('Class Name', axis=1)  # Remove a coluna Class Name pois já está no título
        st.dataframe(df_turma_display)
        
        # Botão de download para a turma específica
        def to_excel_turma(df, nome_turma):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Criar uma cópia do dataframe para manipulação
                df_export = df.copy()
                
                # Adicionar coluna vazia para mesclar
                df_export.insert(0, 'Class_Name_Merge', '')
                
                # Preencher apenas a primeira linha com o nome da turma
                df_export.at[0, 'Class_Name_Merge'] = nome_turma
                
                # Remover a coluna original Class Name
                df_export = df_export.drop('Class Name', axis=1)
                
                # Renomear a coluna mesclada
                df_export = df_export.rename(columns={'Class_Name_Merge': 'Class Name'})
                
                # Escrever no Excel
                df_export.to_excel(writer, index=False, sheet_name=nome_turma[:31])  # Limite de 31 caracteres para nome da aba
                
                # Obter a planilha para mesclar células
                worksheet = writer.sheets[nome_turma[:31]]
                
                # Formatar cabeçalho com cor azul escuro
                for col in range(1, len(df_export.columns) + 1):
                    header_cell = worksheet.cell(row=1, column=col)
                    header_cell.fill = PatternFill(start_color='405ddb', fill_type='solid')
                    header_cell.alignment = Alignment(horizontal='center', vertical='center')
                    header_cell.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
                
                # Lista de cores para as turmas (mesma lista da função principal)
                cores_turmas = [
                    'FFE6E6',  # Rosa claro
                    'E6F3FF',  # Azul claro
                    'E6FFE6',  # Verde claro
                    'FFF0E6',  # Laranja claro
                    'F0E6FF',  # Roxo claro
                    'FFFFE6',  # Amarelo claro
                    'E6FFFF',  # Ciano claro
                    'FFE6F0',  # Rosa pálido
                    'F0FFE6',  # Verde pálido
                    'E6E6FF'   # Azul pálido
                ]
                
                # Obter todas as turmas para mapear cores
                todas_turmas = dfFinal['Class Name'].unique()
                mapeamento_cores = {turma: cores_turmas[i % len(cores_turmas)] for i, turma in enumerate(todas_turmas)}
                
                # Mesclar todas as células da coluna Class Name (já que é só uma turma)
                if len(df_export) > 1:
                    worksheet.merge_cells(f'A2:A{len(df_export) + 1}')
                    # Centralizar verticalmente e horizontalmente
                    worksheet['A2'].alignment = Alignment(horizontal='center', vertical='center')
                    # Aplicar cor da turma
                    if nome_turma in mapeamento_cores:
                        worksheet['A2'].fill = PatternFill(start_color=mapeamento_cores[nome_turma], fill_type='solid')
                
                # Definir borda preta
                black_border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                
                # Centralizar todas as células de dados e aplicar formatação condicional
                acc_total_col = None
                name_col = None
                acc_personalizada_col = None
                for col_idx, col_name in enumerate(df_export.columns, 1):
                    if col_name == 'ACC Total':
                        acc_total_col = col_idx
                    elif col_name == 'Name':
                        name_col = col_idx
                    elif col_name.startswith('ACC Total por') and col_name.endswith('%'):
                        acc_personalizada_col = col_idx
                
                # Obter a cor da turma atual
                cor_da_turma = mapeamento_cores.get(nome_turma, 'FFFFFF')  # Branco como padrão
                
                for row in range(2, len(df_export) + 2):  # Começar da linha 2 (após cabeçalho)
                    for col in range(1, len(df_export.columns) + 1):  # Todas as colunas
                        cell = worksheet.cell(row=row, column=col)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Aplicar cor da turma em todas as células da linha (exceto colunas especiais)
                        if col != acc_total_col and col != acc_personalizada_col:
                            cell.fill = PatternFill(start_color=cor_da_turma, fill_type='solid')
                            cell.border = black_border
                        
                        
                        # Aplicar formatação condicional para ACC Total
                        if col == acc_total_col and cell.value is not None:
                            try:
                                acc_value = float(cell.value)
                                if acc_value < 60:
                                    cell.fill = PatternFill(start_color='d46161', fill_type='solid')
                                    cell.border = black_border
                                if acc_value >= 60:
                                    cell.fill = PatternFill(start_color='73c56c', fill_type='solid')
                                    cell.border = black_border
                            except (ValueError, TypeError):
                                pass  # Ignorar se não conseguir converter para float
                        
                        # Aplicar formatação condicional para ACC Total por % baseada no novo input
                        if col == acc_personalizada_col and cell.value is not None and novo_input > 0:
                            try:
                                acc_personalizada_value = float(cell.value)
                                if acc_personalizada_value >= novo_input:
                                    cell.fill = PatternFill(start_color='73c56c', fill_type='solid')
                                    cell.border = black_border
                                else:
                                    cell.fill = PatternFill(start_color='d46161', fill_type='solid')
                                    cell.border = black_border
                            except (ValueError, TypeError):
                                pass  # Ignorar se não conseguir converter para float
                        
                
                # Ajustar largura das colunas
                worksheet.column_dimensions['A'].width = 90
                worksheet.column_dimensions['B'].width = 50
                worksheet.column_dimensions['C'].width = 15
                worksheet.column_dimensions['D'].width = 40
                worksheet.column_dimensions['E'].width = 40 
                worksheet.column_dimensions['F'].width = 40
                worksheet.column_dimensions['G'].width = 40
                worksheet.column_dimensions['H'].width = 40
                worksheet.column_dimensions['I'].width = 40
                worksheet.column_dimensions['J'].width = 40
                worksheet.column_dimensions['K'].width = 40
                worksheet.column_dimensions['L'].width = 40
                worksheet.column_dimensions['M'].width = 40
                worksheet.column_dimensions['N'].width = 40
                worksheet.column_dimensions['O'].width = 40
                worksheet.column_dimensions['P'].width = 40
                worksheet.column_dimensions['Q'].width = 40
                worksheet.column_dimensions['R'].width = 40
                worksheet.column_dimensions['S'].width = 40
                
            return output.getvalue()
        
        # Nome do arquivo baseado na turma (removendo caracteres especiais)
        nome_arquivo = turma.replace(' ', '_').replace('-', '_').replace('º', '').replace('°', '')
        excel_data_turma = to_excel_turma(df_turma, turma)
        
        st.download_button(
            label=f"📥 Baixar {turma} (XLSX)",
            data=excel_data_turma,
            file_name=f"{nome_arquivo}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"download_{turma}"  # Chave única para cada botão
        )
        st.write("---")
